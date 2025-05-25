import yfinance as yf
import numpy as np
import matplotlib.pyplot as plt
import datetime as datetime
import pandas as pd
from scipy.stats import norm

# Get stock price BHP on 16 May 2025
ticker = "BHP"
stock = yf.Ticker(ticker)
#16 May Historical Data
target_date = "2025-05-16"
hist = stock.history(start="2025-05-16", end="2025-05-17")

#Closing price extract
if hist.empty:
    raise ValueError(f"No stock data found for {ticker} on {target_date}.")
S0 = hist["Close"].iloc[0]  # Closing price on 16 May 2025
print(f"Closing price of {ticker} on {target_date}: {S0}")

# Calculate K 
K = 0.98*S0
print(f"Strike price K calculated as 98% of S0: {K}")

# Calculate T 
Valuation_Date = datetime.datetime(2025, 5, 16)
Expiration_Date = datetime.datetime(2027, 9, 15)
T = (Expiration_Date - Valuation_Date).days / 365
print(f"Time to expiration (T) in years: {T}")

# Pull volatility (sigma) from cell S15 in sheet "Worksheet"
vol_file = "Rates_Volatilities.xlsx"
from openpyxl import load_workbook

# Open the workbook and access the value from R12
wb = load_workbook(filename=vol_file, data_only=True)
ws = wb["BHP"]
sigma = ws["S15"].value  # This returns the evaluated result (not the formula)

if sigma is None:
    raise ValueError("Volatility (sigma) in cell S15 returned None — please check if Excel formula has been calculated and saved.")

# If sigma looks like a percentage (e.g., 20.24), convert to decimal
if sigma > 1:
    sigma = sigma / 100

print(f"Volatility (sigma): {sigma:.2%}")

#Calculate R
df = pd.read_excel('Risk Free - Yield Curve.xlsx')
rate_2y = df.loc[df['Tenor'] == '2Y', 'I1 Mid YTM AUD Australia Sovereign Curve 05/16/25 Yield'].values[0]
# Convert percentage to decimal if needed
r = rate_2y_decimal = rate_2y / 100 if rate_2y > 1 else rate_2y

print(f"2-Year Risk-Free Rate (decimal): {rate_2y_decimal}")

from instrument_classes import EuropeanCall
EC = EuropeanCall(S0, K, r, T, sigma)
price = EC.price()
print(f"Call Option Price: {price}")